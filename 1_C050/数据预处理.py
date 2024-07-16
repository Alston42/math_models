import numpy as np
from numpy import arange
from matplotlib import pyplot
from pandas import read_csv
from pandas.plotting import scatter_matrix 
import pandas as pd
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.model_selection import KFold
from sklearn.model_selection import cross_val_score
from sklearn.model_selection import GridSearchCV
from sklearn.linear_model import LinearRegression
from sklearn.linear_model import Lasso
from sklearn.linear_model import ElasticNet
from sklearn.tree import DecisionTreeRegressor
from sklearn.neighbors import KNeighborsRegressor
from sklearn.svm import SVR
from sklearn.pipeline import Pipeline
from sklearn.ensemble import RandomForestRegressor
from sklearn.ensemble import GradientBoostingRegressor
from sklearn.ensemble import ExtraTreesRegressor
from sklearn.ensemble import AdaBoostRegressor
from sklearn.metrics import mean_squared_error
from sklearn.ensemble import RandomForestRegressor
from scipy.stats import skew, kurtosis
from scipy.stats import spearmanr
from sklearn.cluster import KMeans
import openpyxl
import pprint
from scipy.stats import skew
from matplotlib import font_manager
from matplotlib import pyplot as plt

font_path = 'C:/Windows/Fonts/simhei.ttf'  # 中文字体
font_manager.fontManager.addfont(font_path)
plt.rcParams['font.family'] = font_manager.FontProperties(fname=font_path).get_name()



wb =  openpyxl.load_workbook('D:/A LBYYY/数学建模/~数学建模算法学习/2023 C题/table1.xlsx')
sheet = wb['Sheet1']
data = sheet.values
columns = next(data) 
df = pd.DataFrame(data, columns=columns)
df = df.dropna() #空白行
print(df.describe()) #统计性描述
print(df.corr(method='pearson'))

spearman_corr, spearman_pval = spearmanr(df)
fig=pyplot.figure()
ax=fig.add_subplot(111)
cax = ax.matshow(df.corr(),vmin=-1,vmax=1,interpolation='none')

for i in range(spearman_corr.shape[0]):
    for j in range(spearman_corr.shape[1]):
        ax.text(x=j, y=i, s="{:.2f}\n(p={:.2f})".format(spearman_corr[i, j], spearman_pval[i, j]),
                va='center', ha='center', color='black', fontsize=8)

fig.colorbar(cax)
ticks = np.arange(0, 5 ,1)  #这里要改参数 x,y轴起始值终止值和步长
ax.set_xticks(ticks)
ax.set_yticks(ticks)
ax.set_xticklabels(columns)
ax.set_yticklabels(columns)
pyplot.show()
print("Spearman correlation coefficients:")
print(spearman_corr)
print("\nSpearman p-values:")
print(spearman_pval)

wb1 =  openpyxl.load_workbook('D:/A LBYYY/数学建模/~数学建模算法学习/2023 C题/table2.xlsx')
sheet1 = wb1['Sheet1']
data1 = sheet1.values
columns1 = next(data1) 
df1 = pd.DataFrame(data1, columns=columns1)
description = df1.describe()
table3 = openpyxl.Workbook()
sheet3 = table3.active


for c, col_name in enumerate([''] + list(description.columns), start=1):
    sheet3.cell(row=1, column=c, value=col_name)

for r, row in enumerate(description.itertuples(), start=2):
    sheet3.cell(row=r, column=1, value=row[0])  
    for c, value in enumerate(row[1:], start=2):
        sheet3.cell(row=r, column=c, value=value)


table3.save('D:/A LBYYY/数学建模/~数学建模算法学习/2023 C题/table3.xlsx')

